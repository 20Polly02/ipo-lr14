from django.shortcuts import render
import json
from django.http import JsonResponse #новое
from django.contrib.auth.decorators import login_required # новое
from django.views.decorators.http import require_POST #новое
from django.contrib.auth import login, authenticate #новое
from .forms import UserRegistrationForm, LoginForm #новое
from django.urls import reverse
from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO
from django.core.mail import send_mail
from django.conf import settings
from django.core.mail import EmailMessage
from openpyxl.styles import Font
from .models import Product,Category,Maker,Basket,Basket_elem,Order,OrderItem
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import Filter

from .serializers import ProductModelSerializer,CategoryModelSerializer,MakerModelSerializer,BasketModelSerializer,Basket_elementModelSerializer,OrderModelSerializer,OrderItemModelSerializer
from rest_framework import viewsets
def home_page(request):
    return render(request,'home_page.html')

def about_autor(request):
    return render(request,'about_autor.html')

def home_goods_store(request):
    return render(request,'home_goods_store.html')
def speciality(request):
    specialities=[]
    with open("data/dump.json", 'r', encoding='utf-8') as file:  
        info_file = json.load(file)
    for specialty in info_file:  
        if specialty.get("model") == "data.specialty": 
            speciality = {
                "code": specialty["fields"].get("code"),
                "pk": specialty.get("pk"), 
                "title" : specialty["fields"].get("title"),  
                "c_type": specialty["fields"].get("c_type")  
            }
            specialities.append(speciality)
    return render(request,'speciality.html',{'specialities':specialities})
def speciality_found(request):
    c = request.GET.get('c')
    speciality = []  
    with open("data/dump.json", 'r', encoding='utf-8') as file:
        read_file = json.load(file)
        for specialty in read_file:
            if specialty.get("model") == "data.specialty":
                if specialty["fields"].get("code") == c:
                    specialty_f = {
                        "code": specialty["fields"].get("code"),
                        "pk": specialty.get("pk"),
                        "title": specialty["fields"].get("title"),
                        "educational": specialty["fields"].get("c_type"),
                    }
                    speciality.append(specialty_f)  

    return render(request, "speciality_found.html", {'speciality': speciality})

def speciality_id(request,id):
    speciality = []  
    with open("data/dump.json", 'r', encoding='utf-8') as file:
        read_file = json.load(file)
        for specialty in read_file:
            if specialty.get("model") == "data.specialty":
                if int(specialty.get("pk")) == int(id):
                    specialty_i= {
                        "code": specialty["fields"].get("code"),
                        "pk": specialty.get("pk"),
                        "title": specialty["fields"].get("title"),
                        "educational": specialty["fields"].get("c_type"),
                    }
                    speciality.append(specialty_i)  

    return render(request, "speciality_found.html", {'speciality': speciality}) 
def product_list(request):
    products = Product.objects.all()
    search_query = request.GET.get('search', '')
    
    # Применяем поиск, если есть запрос
    if search_query:
        products = products.filter(name__icontains=search_query)
    
    # Применяем другие фильтры
    form = Filter(request.GET or None)
    if form.is_valid():
        if form.cleaned_data['category']:
            products = products.filter(category__name__iexact=form.cleaned_data['category'])
        if form.cleaned_data['maker']:
            products = products.filter(maker__name__iexact=form.cleaned_data['maker'])
    
    return render(request, 'shop/product_list.html', {
        'products': products,
        'form': form,
        'categories': Category.objects.all(),  # Добавьте это
        'makers': Maker.objects.all()         # И это
    })
def product_detail(request,id):
    try:
        product = Product.objects.get(id=id)
    except Product.DoesNotExist:
        messages.error(request, "Товар не найден.")
    return render(request, 'shop/product_detail.html', {'product': product})

@login_required

def add_product_cart(request, product_id):
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist or  UnboundLocalError :
        messages.error(request, "Товар не найден.")
        return redirect('cart_user')
    basket, created = Basket.objects.get_or_create(user=request.user)
    quantity = int(request.POST.get('quantity', 1))

    basket_element, created = Basket_elem.objects.get_or_create(
        basket=basket,
        product=product,
        defaults={'quantity': quantity}
    )
    if not created:
        basket_element.quantity += quantity
        basket_element.save()
        messages.success(request, "Количество товара обновлено в корзине.")
    else:
        messages.success(request, "Товар добавлен в корзину.")
    return redirect('cart_user')

@login_required
@require_POST

def update_cart(request, item_id):
    try:
        basket_element = Basket_elem.objects.get(id=item_id, basket__user=request.user)
    except Basket_elem.DoesNotExist:
        messages.error(request, "Не удалось обновить корзину")
        return redirect('cart_user')

    try:
        quantity = int(request.POST.get('quantity', 1))
    except (ValueError, TypeError):
        quantity = 1

    if quantity <= 0:
        basket_element.delete()
        messages.success(request, "Товар удалён из корзины")
    elif quantity <= basket_element.product.quantity_in_stock:
        basket_element.quantity = quantity
        basket_element.save()
        messages.success(request, "Количество товара обновлено")
    else:
        messages.error(request, "Недостаточно товара на складе")

    return redirect('cart_user')

@login_required

def remove_from_cart(request, item_id):
    try:
        basket_element = Basket_elem.objects.get(id=item_id, basket__user=request.user)
        basket_element.delete()
        messages.success(request, "Товар удалён из корзины")
    except Basket_elem.DoesNotExist:
        messages.error(request, "Товар не найден в корзине")
    return redirect('cart_user')
    
@login_required

def cart_user(request):
    basket, created = Basket.objects.get_or_create(user=request.user)
    basket_elements = basket.basket_elem.select_related('product').all()
    total = sum(item.element_cost() for item in basket_elements)
    context = {
        'cart_items': basket_elements,
        'total': total,
    }
    return render(request, 'shop/cart.html', context)

def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('cart_user')
    else:
        form = UserRegistrationForm()
    return render(request, 'register.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('cart_user')
    else:
        form = LoginForm()
    return render(request, 'login_user.html', {'form': form})

@login_required

def checkout(request):
    try:
      
        cart = Basket.objects.get(user=request.user)
        cart_items = Basket_elem.objects.filter(basket=cart)
        
        if not cart_items.exists():
            return redirect('cart_view')
        
        total_price = sum(item.product.price * item.quantity for item in cart_items)
        
        if request.method == 'POST':
  
                order = Order.objects.create(
                    user=request.user,
                    total_price=total_price,

                )
                
                OrderItem.objects.bulk_create([
                    OrderItem(
                        order=order,
                        product=item.product,
                        quantity=item.quantity,
                        price=item.product.price
                    ) for item in cart_items
                ])
                
                # Генерируем Excel-чек
                excel_file = generate_excel_receipt(order)
                
                # Отправляем email с чеком
                send_mail(request.user.email, order, excel_file, total_price)
                
                cart_items.delete()
                
                messages.success(request, 'Заказ успешно оформлен! Чек отправлен на вашу почту.')
                return redirect('cart_user')
        
        return render(request, 'shop/checkout.html', {
            'cart_items': cart_items,
            'total_price': total_price
        })
        
    except Basket.DoesNotExist:
        messages.warning(request, "Корзина не найдена!")
        return redirect('cart_user')

def generate_excel_receipt(order):
    """Генерирует Excel-чек и возвращает BytesIO объект"""
    wb = Workbook() 
    ws = wb.active 
    ws.title = "Чек заказа" 
    
   
    bold_font = Font(bold=True) 
    
    headers = ["№", "Товар", "Количество", "Цена", "Сумма"] 
    ws.append(headers) 
    for cell in ws[1]:
        cell.font = bold_font 

    for idx, item in enumerate(order.items.all(), start=1):
      
        ws.append([
            idx,
            item.product.name,
            item.quantity,
            item.price,
            item.price * item.quantity
        ])
    
    ws.append(["", "", "", "Итого:", order.total_price])
    for cell in ws[ws.max_row]:
        cell.font = bold_font
    buffer = BytesIO() 
    wb.save(buffer) 
    buffer.seek(0)

def send_mail(email, order, excel_file, total_price):

    email_msg = EmailMessage(
        subject=f'Чек заказа №{order.id} от {order.created.strftime("%d.%m.%Y")}', 
        body=f'''
        Благодарим за ваш заказ!
        
        Номер заказа: {order.id}
        Дата: {order.created.strftime("%d.%m.%Y %H:%M")}
        Сумма заказа: {total_price} руб.
        
        Детали заказа в приложенном файле.
        ''',
        from_email=settings.DEFAULT_FROM_EMAIL, 
        to=[email],
    )
    

    email_msg.attach(
        filename=f'order_{order.id}_receipt.xlsx', 
        content=excel_file.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
       
    )
    
    email_msg.send()

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all() 
    serializer_class = ProductModelSerializer  

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategoryModelSerializer

class MakerViewSet(viewsets.ModelViewSet):
    queryset = Maker.objects.all() 
    serializer_class = MakerModelSerializer 

class BasketViewSet(viewsets.ModelViewSet):
    queryset = Basket.objects.all()
    serializer_class = BasketModelSerializer

class Basket_elementViewSet(viewsets.ModelViewSet):
    queryset = Basket_elem.objects.all() 
    serializer_class = Basket_elementModelSerializer  

class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderModelSerializer

class OrderItemViewSet(viewsets.ModelViewSet):
    queryset = OrderItem.objects.all()
    serializer_class = OrderItemModelSerializer
# Create your views here.
